# -*- coding:UTF-8 -*-
import copy
import pandas as pd
import openpyxl
from openpyxl.styles import Font

num2month = ("", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")

class CwwSinker:
    '''
    根据cww个人记账习惯的 sinker
    - 分为两个队列(daily/huge)
    - 自动按月拆分写进多个sheet(目前必须从1月开始)
    '''
    def __init__(self, ctx) -> None:
        self.sink_len = 11
        self.alias = {
            "收款方备注:二维码收款": "二维码收款",
            "高德地图打车订单": "taxi",
            "顺丰速运有限公司": "顺丰快递",
        }
        if ctx.conf_dict.get("sinker", {}).get("sinker_alias", None) is not None:
            self.alias.update(ctx.conf_dict.get("sinker", {}).get("sinker_alias", {}))
        ctx.info_log.write("sinker_alias", len(self.alias))
        self.font = Font('等线')

    def prepare_df(self, ctx):
        for record in ctx.bill_records:
            record.oppo_account = self.alias.get(record.oppo_account, record.oppo_account)
            record.description = self.alias.get(record.description, record.description)
            record.description = f"{record.oppo_account[:self.sink_len]}:{record.description[:self.sink_len]}"
        df = pd.DataFrame(ctx.bill_records)
        df = df.drop(columns=["category", "source", "oppo_account"])
        # df["date"] = df["date"].dt.strftime('%d')
        # df["date"] = df["date"].dt.day
        return df

    def find_row_for_huge_item(self, daily_list, huge_item):
        for daily_item in daily_list:
            if daily_item[0] == huge_item[0] and daily_item[3] == "":
                daily_item[3], daily_item[4] = huge_item[1], huge_item[2]
                return daily_item
        # 满了 or 当天没有: 在 next_day 前 insert
        for daily_item in daily_list:
            if  huge_item[0]+1 <= daily_item[0]:
                idx = daily_list.index(daily_item)
                daily_list.insert(idx, huge_item[0:1] + ["", ""] + huge_item[1:3])
                return daily_list[idx-1]
        raise Exception(f"[CwwSinker] sth wrong in find_row_for_huge_item: {huge_item}")
    def merge_queue(self, ctx, daily_lists, huge_lists, _csv="data/tmp"):
        assert len(daily_lists) == len(huge_lists)
        merged_lists = []
        for i, (daily_list, huge_list) in enumerate(zip(daily_lists, huge_lists)):
            month = daily_list[0][3]
            assert i + 1 == month  # 从一月开始计算
            daily_list = [x[:3] + ["", ""] for x in daily_list]
            daily_list.append([32, "", "", "", ""])  # 虚拟日期, idx用
            for huge_item in huge_list:
                item = self.find_row_for_huge_item(daily_list, huge_item)
            daily_list.pop()  # 删除虚拟日期
            merged_lists.append(daily_list)
            if _csv != "":
                df_merge = pd.DataFrame(daily_list, columns=["date", "description", "amount", "logic1", "logic2"])
                df_merge.to_csv(f"{_csv}/tmp_merge_{month:02d}.csv", index=False)
                # ctx.info_log.write(f"merged_csv", f"{_csv}/tmp_merge.csv")
        return merged_lists

    def fetch(self, ctx):
        df = self.prepare_df(ctx)
        ctx.info_log.write("begin_sink", df.shape[0])
        daily_lists = self.sink_queue(ctx, df, "daily")
        huge_lists = self.sink_queue(ctx, df, "huge")
        merged_lists = self.merge_queue(ctx, daily_lists, huge_lists)
        self.sink_excel(ctx, merged_lists)
        
        ctx.info_log.flush("CwwSinker")

    def sink_queue(self, ctx, df, _queue, _csv="data/tmp"):
        df_queue = df[df["queue"]==_queue].reset_index(drop=True)
        df_queue = df_queue.drop(columns=["queue"])

        # for 显示清爽
        df_queue['amount'] = df_queue['amount'].astype(int)
        ctx.info_log.write(f"drop_{_queue}_tiny", df_queue[abs(df_queue["amount"])<=1].shape[0]);
        df_queue = df_queue.drop(df_queue[abs(df_queue["amount"]) <= 1].index)

        if _csv != "":
            df_queue.to_csv(f"{_csv}/tmp_{_queue}.csv", index=False)
            # ctx.info_log.write(f"{_queue}_csv", f"{_csv}/tmp_{_queue}.csv")
        ctx.info_log.write(f"sink_{_queue}", df_queue.shape[0])
        
        # 按月拆分
        df_queue["month"] = df_queue["date"].dt.month
        df_queue["date"] = df_queue["date"].dt.day
        grouped = df_queue.groupby('month')
        df_queue_month = [grouped.get_group(x).values.tolist() for x in grouped.groups]
        return df_queue_month

    def sink_excel(self, ctx, merged_lists):
        wb = openpyxl.Workbook()
        ws_year = wb.worksheets[0]
        ws_year.append(["month", "daily", "huge", "salary", "monthly"])
        for i, month_list in enumerate(merged_lists):
            month = num2month[i + 1]
            ws = wb.create_sheet(month)
            for i in range(len(month_list)-1, 0, -1):
                if month_list[i][0] ==  month_list[i-1][0]:
                    month_list[i][0] = ""
            salary = 0
            for row in month_list:
                ws.append(row)
                salary += row[4] if row[3]==":工资" else 0

            # 月度汇总
            ws['F2'], ws['F3'] = "daily", f'=SUM(C:C)'
            ws['G2'], ws['G3'] = "huge", f'=SUM(E:E)-H3'
            ws['H2'], ws['H3'] = "salary", salary
            ws['I2'], ws['I3'] = "total", f'=SUM(F3:H3)'
            ws_year.append([month, f'={month}!F3', f'={month}!G3', f'={month}!H3', f'={month}!I3'])
            # 格式刷
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['D'].width = 17
            for row in range(ws.max_row):
                for column in range(ws.max_column):
                    ws.cell(row=row+1,column=column+1).font = self.font
        ws_year.append(["sum", "", "", "", f'=SUM(E2:E{ws_year.max_row})'])
        wb.save(f"data/cww账单{ctx.year}.xlsx")
        # import pdb; pdb.set_trace()
